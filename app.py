import concurrent.futures
import json
import os
import re
import secrets
from datetime import datetime, timedelta

import pandas as pd
from fastapi import (BackgroundTasks, Depends, FastAPI, File, HTTPException,
                     UploadFile)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi_mail import ConnectionConfig, FastMail, MessageSchema
from fastapi import FastAPI, Request, HTTPException
from jose import jwt, JWTError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pydantic import BaseModel, EmailStr

from models import LoginRequest, NewPasswordRequest, ResetPasswordRequest
from utils import clean_phone_number, query_cnam_api

SECRET_KEY = secrets.token_urlsafe(32)

USER_FILE = "users.json"

# Configure email settings
conf = ConnectionConfig(
    MAIL_USERNAME="c.hardik125@gmail.com",
    MAIL_PASSWORD="frzdeegakgsbluqu",
    MAIL_FROM="c.hardik125@gmail.com",
    MAIL_PORT=587,
    MAIL_SERVER="smtp.gmail.com",
    MAIL_STARTTLS=True,
    MAIL_SSL_TLS=False,
    USE_CREDENTIALS=True,
    VALIDATE_CERTS=False,
)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "*"
    ],  # Replace "*" with your frontend's URL for stricter rules, e.g., ["http://localhost:3000"]
    allow_credentials=True,
    allow_methods=["*"],  # Use specific methods like ["GET", "POST"] if needed
    allow_headers=["*"],  # Specify headers if needed
)


UPLOAD_FOLDER = "./uploads"


def clean_name(name):
    return re.findall(r"\w+", name.upper())


@app.post("/upload")
async def create_upload_file(file: UploadFile = File(...)):
    if file.filename.split(".")[-1].lower() != "xlsx":
        return JSONResponse(
            content={"error": "Only xlsx files are allowed"}, status_code=400
        )

    # Remove any existing files in the upload folder
    for existing_file in os.listdir(UPLOAD_FOLDER):
        file_path = os.path.join(UPLOAD_FOLDER, existing_file)
        if os.path.isfile(file_path):
            os.remove(file_path)

    # Save the file to the upload folder
    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    with open(file_path, "wb") as f:
        f.write(file.file.read())

    # Simulate long processing time
    await process_excel(file_path)

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="processed_file.xlsx",
    )


async def process_excel(file_path):
    process_file(file_path)
    return "Processing complete"


def process_file(file_path):
    df = pd.read_excel(file_path)

    valid_headers = ["phone n", "phonen"]
    headers = [col for col in df.columns]
    phone_columns = [
        col
        for col in headers
        if re.fullmatch(r"(Relative\s?\d*\s?)?[Pp]hone\s?\d+$", col)
    ]

    new_column = ["", ""]
    for phone_column in phone_columns:
        new_column.append(phone_column + "No")
        new_column.append(phone_column + "API Name")

    new_column.append("")
    new_column.append("")

    for phone_column in phone_columns:
        new_column.append(phone_column + "No")
        new_column.append(phone_column + "API Name")

    updated_columns = df.columns.tolist() + new_column
    df = df.reindex(columns=updated_columns)
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    sheet = wb.active

    relative_values = list()
    for phone_column in phone_columns:

        relative_value = re.split(r"(?<=\d) ", phone_column)[0]
        relative_values.append(relative_value)

    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        futures = []
        for index, row in df.iterrows():
            for int_idx, column in enumerate(phone_columns):

                phone_number = row[column]
                if isinstance(phone_number, float) and phone_number.is_integer():
                    phone_number = str(int(phone_number))
                elif isinstance(phone_number, str) and phone_number.endswith(".0"):
                    phone_number = phone_number.rstrip(".0")

                future = executor.submit(
                    query_cnam_api, clean_phone_number(str(phone_number))
                )
                futures.append((index, row, int_idx, future))

        results = []
        for index, row, int_idx, future in futures:
            result = future.result()
            results.append((index, row, int_idx, result))

        for index, row, int_idx, result in results:
            api_name = result.get("name", "").upper()
            api_name_parts = clean_name(api_name)

            # for relative_value in relative_values:
            first_name_column = (
                f"{relative_values[int_idx]} First Name"
                if "Relative" in relative_values[int_idx]
                else "First Name"
            )
            last_name_column = (
                f"{relative_values[int_idx]} Last Name"
                if "Relative" in relative_values[int_idx]
                else "Last Name"
            )
            excel_name_parts = clean_name(
                f"{row[first_name_column]} {row[last_name_column]}"
            )

            phone_number = row[phone_columns[int_idx]]
            clean_number = clean_phone_number(phone_number)

            is_match = any(api_part in excel_name_parts for api_part in api_name_parts)
            light_green_fill = PatternFill(
                start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
            )
            light_red_fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )

            if is_match:
                col_index = len(row) - (len(phone_columns) * 4) + (int_idx * 2) - 1
                sheet.cell(index + 2, col_index).value = clean_number
                sheet.cell(index + 2, col_index + 1).value = api_name
                sheet.cell(index + 2, col_index + 1).fill = light_green_fill
            elif not is_match:
                col_index = len(row) - (len(phone_columns) * 2) + (int_idx * 2) + 1
                sheet.cell(index + 2, col_index).value = clean_number
                sheet.cell(index + 2, col_index + 1).value = api_name
                sheet.cell(index + 2, col_index + 1).fill = light_red_fill

    wb.save(file_path)

    return "processing complete"


def create_reset_token(email: str):
    payload = {
        "email": email,
        "exp": datetime.utcnow() + timedelta(hours=1),  # Token expires in 1 hour
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")


def get_users():
    with open(USER_FILE, "r") as file:
        users = json.load(file)
    return users


@app.post("/login")
async def login(request: LoginRequest):
    users = get_users()
    user = next(
        (
            u
            for u in users
            if u.get("email") == request.email and u.get("password") == request.password
        ),
        None,
    )

    if user:
        # Create a JWS token
        payload = {
            "email": user["email"],
            "exp": datetime.utcnow() + timedelta(hours=1),
        }
        token = jwt.encode(payload, SECRET_KEY, algorithm="HS256")

        return {"message": "Login successful", "token": token, "email": user["email"]}
    else:
        raise HTTPException(status_code=404, detail="No user found")


@app.post("/send_reset_link")
async def send_reset_link(
    request: ResetPasswordRequest, background_tasks: BackgroundTasks
):
    users = get_users()
    user = next((u for u in users if u.get("email") == request.email), None)

    if user:
        token = create_reset_token(request.email)
        link = f"http://localhost:3000/reset-password?token={token}"
        subject = "Password Reset"
        body = (
            f"Hi,\n\nClick the link below to reset your password:\n{link}\n\nThank you."
        )

        message = MessageSchema(
            subject=subject,
            recipients=[request.email],
            body=body,
            subtype="plain",
        )

        fm = FastMail(conf)
        background_tasks.add_task(fm.send_message, message)
        return {"message": "Password reset link sent successfully","token":token}
    else:
        raise HTTPException(status_code=404, detail="No user found")


@app.post("/reset_password")
async def reset_password(request: Request, body: NewPasswordRequest):
    try:
        # Get the Authorization header
        auth_header = request.headers.get("Authorization")
        
        if not auth_header:
            raise HTTPException(status_code=401, detail="Authorization header missing")
        
        # Ensure the header starts with "Bearer "
        if not auth_header.startswith("Bearer "):
            raise HTTPException(status_code=401, detail="Invalid Authorization header format")
        
        # Extract the token
        token = auth_header.replace("Bearer ", "").strip()

        # Decode the JWT token
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        email = payload.get("email")  # Extract email from the payload

        if not email:
            raise HTTPException(status_code=403, detail="Invalid token: email not found in payload")

        # Get the list of users
        users = get_users()
        user = next((u for u in users if u.get("email") == email), None)

        if user:
            # Update the user's password
            user["password"] = body.new_password
            with open(USER_FILE, "w") as file:
                json.dump(users, file, indent=4)
            return {"message": "Password reset successfully"}
        else:
            raise HTTPException(status_code=404, detail="No user found")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except JWTError as e:
        raise HTTPException(status_code=401, detail=f"Invalid token: {str(e)}")
