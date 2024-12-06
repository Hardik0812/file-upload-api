import os
import re

import pandas as pd
from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook

from utils import clean_phone_number, query_cnam_api

app = FastAPI()

# Define the upload folder
UPLOAD_FOLDER = './uploads'
# Create the upload folder if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def clean_name(name):
    return re.findall(r"\w+", name.upper())

def process_file():
    file_path = os.path.join(UPLOAD_FOLDER, "source1.xlsx")
    df = pd.read_excel(file_path)
    # new_row = pd.Series([None] * len(df.columns), index=df.columns)

    # # Insert the new blank row at the top
    # df.insert(0, new_row.name, new_row)

    valid_headers = ["phone n", "phonen"]
    headers = [col for col in df.columns]
    phone_columns = [
        col
        for col in headers
        if re.fullmatch(r"(Relative ?\d+ )?Phone\d+$", col, re.IGNORECASE)
    ]
    new_column = ["",""]
    for phone_column in phone_columns:
        new_column.append(phone_column + "No")
        new_column.append(phone_column + "API Name")

    new_column.append("")
    new_column.append("")

    for phone_column in phone_columns:
        new_column.append(phone_column + "No")
        new_column.append(phone_column + "API Name")

    updated_columns = df.columns.tolist() + new_column
    df = df.reindex(columns=updated_columns)
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    sheet = wb.active

    for index, row in df.iterrows():
        print(row)
        for int_idx, column in enumerate(phone_columns):
            # phone_num = next((row[col] for col in row.index if col == phone_column), None)
            phone_number = row[column]
            clean_number = clean_phone_number(str(phone_number))
            api_response = query_cnam_api(clean_number)
            print(api_response)
            api_name = api_response.get("name", "").upper()
            api_name_parts = clean_name(api_name)
            excel_name_parts = clean_name(f"{row['First Name']} {row['Last Name']}")
            is_match = any(
                        api_part in excel_name_parts for api_part in api_name_parts
                    )
                    
            if is_match:
                col_index = len(row)  - 26 + (int_idx *2 )
                sheet.cell(index + 2, col_index).value = clean_number
                sheet.cell(index + 2, col_index + 1).value = api_name
                # df.at[2, col_index + 1] = "yes"
            else:
                col_index = len(row)  - 9 + (int_idx *2 )
                sheet.cell(index + 2, col_index).value = clean_number
                sheet.cell(index + 2, col_index + 1).value = api_name


    wb.save(file_path)
    

    if not all(header in valid_headers for header in headers):
        return JSONResponse(content={"error": "Invalid file headers"}, status_code=400)



@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    # Check if the file is an xlsx file
    if file.filename.split('.')[-1].lower() != 'xlsx':
        return JSONResponse(content={"error": "Only xlsx files are allowed"}, status_code=400)
    
    # Save the file to the upload folder
    with open(os.path.join(UPLOAD_FOLDER, file.filename), "wb") as f:
        f.write(file.file.read())
    process_file()
    return JSONResponse(content={"message": "File uploaded successfully"}, status_code=200)


    